from flask import Blueprint, render_template, request, jsonify, send_file, flash, redirect, url_for
from flask_login import login_required, current_user
from functools import wraps
from models import db, User, Test, TestAttempt, Category
from datetime import datetime, timedelta
from sqlalchemy import func
import io

analytics_bp = Blueprint('analytics', __name__, url_prefix='/analytics')


def analyst_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role not in ('admin', 'hr', 'manager'):
            flash('Недостаточно прав.', 'danger')
            return redirect(url_for('main.dashboard'))
        return f(*args, **kwargs)
    return decorated


@analytics_bp.route('/')
@login_required
@analyst_required
def index():
    # Базовая статистика
    total_users = User.query.filter_by(is_active=True).count()
    total_tests = Test.query.filter_by(is_active=True).count()
    total_attempts = TestAttempt.query.filter(TestAttempt.finished_at.isnot(None)).count()
    passed_attempts = TestAttempt.query.filter(
        TestAttempt.finished_at.isnot(None), TestAttempt.passed == True
    ).count()
    global_pass_rate = round(passed_attempts / total_attempts * 100, 1) if total_attempts else 0

    # Активность за последние 30 дней
    since = datetime.utcnow() - timedelta(days=30)
    recent_attempts = TestAttempt.query.filter(
        TestAttempt.finished_at >= since,
        TestAttempt.finished_at.isnot(None)
    ).count()

    # Топ проблемные тесты (низкий процент сдачи)
    problem_tests = []
    for test in Test.query.filter_by(is_active=True).all():
        finished = [a for a in test.attempts if a.finished_at]
        if len(finished) >= 3:
            rate = sum(1 for a in finished if a.passed) / len(finished) * 100
            problem_tests.append((test, round(rate, 1), len(finished)))
    problem_tests.sort(key=lambda x: x[1])
    problem_tests = problem_tests[:5]

    # Последние попытки
    recent = TestAttempt.query.filter(TestAttempt.finished_at.isnot(None))\
        .order_by(TestAttempt.finished_at.desc()).limit(10).all()

    return render_template('analytics/index.html',
                           total_users=total_users,
                           total_tests=total_tests,
                           total_attempts=total_attempts,
                           global_pass_rate=global_pass_rate,
                           recent_attempts=recent_attempts,
                           problem_tests=problem_tests,
                           recent=recent)


@analytics_bp.route('/api/chart-data')
@login_required
@analyst_required
def chart_data():
    """JSON-данные для графиков на дашборде."""
    days = 30
    since = datetime.utcnow() - timedelta(days=days)

    # Попытки по дням
    daily = []
    for i in range(days):
        day = since + timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0)
        day_end = day_start + timedelta(days=1)
        count = TestAttempt.query.filter(
            TestAttempt.finished_at >= day_start,
            TestAttempt.finished_at < day_end
        ).count()
        daily.append({'date': day_start.strftime('%d.%m'), 'count': count})

    # Результаты по категориям
    categories = Category.query.all()
    cat_data = []
    for cat in categories:
        test_ids = [t.id for t in cat.tests]
        if not test_ids:
            continue
        attempts = TestAttempt.query.filter(
            TestAttempt.test_id.in_(test_ids),
            TestAttempt.finished_at.isnot(None)
        ).all()
        if attempts:
            avg = round(sum(a.score for a in attempts) / len(attempts), 1)
            cat_data.append({'name': cat.name, 'avg_score': avg, 'count': len(attempts)})

    # Распределение баллов
    score_ranges = {'0-39': 0, '40-59': 0, '60-74': 0, '75-89': 0, '90-100': 0}
    all_finished = TestAttempt.query.filter(TestAttempt.finished_at.isnot(None)).all()
    for a in all_finished:
        s = a.score
        if s < 40:
            score_ranges['0-39'] += 1
        elif s < 60:
            score_ranges['40-59'] += 1
        elif s < 75:
            score_ranges['60-74'] += 1
        elif s < 90:
            score_ranges['75-89'] += 1
        else:
            score_ranges['90-100'] += 1

    # По отделам
    dept_data = []
    if current_user.role in ('admin', 'hr'):
        depts = db.session.query(User.department, func.count(User.id))\
            .filter(User.department.isnot(None), User.department != '')\
            .group_by(User.department).all()
        for dept, count in depts:
            dept_users = User.query.filter_by(department=dept).all()
            user_ids = [u.id for u in dept_users]
            dept_attempts = TestAttempt.query.filter(
                TestAttempt.user_id.in_(user_ids),
                TestAttempt.finished_at.isnot(None)
            ).all()
            if dept_attempts:
                avg = round(sum(a.score for a in dept_attempts) / len(dept_attempts), 1)
                rate = round(sum(1 for a in dept_attempts if a.passed) / len(dept_attempts) * 100, 1)
                dept_data.append({'dept': dept, 'avg_score': avg, 'pass_rate': rate, 'count': len(dept_attempts)})

    return jsonify({
        'daily': daily,
        'categories': cat_data,
        'score_distribution': [{'range': k, 'count': v} for k, v in score_ranges.items()],
        'departments': dept_data,
    })


@analytics_bp.route('/export/excel')
@login_required
@analyst_required
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('Библиотека openpyxl не установлена.', 'danger')
        return redirect(url_for('analytics.index'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Результаты тестирования'

    header_fill = PatternFill('solid', fgColor='1E3A5F')
    header_font = Font(color='FFFFFF', bold=True)
    headers = ['ID', 'Сотрудник', 'Отдел', 'Тест', 'Категория',
               'Дата начала', 'Дата завершения', 'Длительность (мин)',
               'Балл (%)', 'Сдан', '№ Попытки']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    attempts = TestAttempt.query.filter(TestAttempt.finished_at.isnot(None))\
        .order_by(TestAttempt.finished_at.desc()).all()

    for a in attempts:
        ws.append([
            a.id,
            a.user.full_name if a.user else '',
            a.user.department if a.user else '',
            a.test.title if a.test else '',
            a.test.category.name if a.test and a.test.category else '',
            a.started_at.strftime('%d.%m.%Y %H:%M') if a.started_at else '',
            a.finished_at.strftime('%d.%m.%Y %H:%M') if a.finished_at else '',
            a.duration_minutes,
            a.score,
            'Да' if a.passed else 'Нет',
            a.attempt_number,
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='ktz_results.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
